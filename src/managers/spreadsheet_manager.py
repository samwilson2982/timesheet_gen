from openpyxl import Workbook
from openpyxl import load_workbook


class SpreadSheetManager(object):

    def __init__(self, filename):
        # start at A0
        self.filename = filename
        self.cursor = (1, 1)
        self.workbook = None
        self.worksheet = None

    def shift_right(self, ammount=1):
        self.cursor = (self.cursor[0] + ammount, self.cursor[1])

    def shift_left(self, ammount=1):
        self.cursor = (self.cursor[0] - ammount, self.cursor[1])

    def shift_up(self, ammount=1):
        self.cursor = (self.cursor[0], self.cursor[1] - ammount)

    def shift_down(self, ammount=1):
        self.cursor = (self.cursor[0], self.cursor[1] + ammount)

    def set_position(self, x, y):
        self.cursor = (x, y)

    def write_value_to_cell(self, val, auto_move=True):
        cell = self.worksheet.cell(column=self.cursor[0], row=self.cursor[1])
        cell.value = val
        if auto_move:
            self.shift_down()

    def write_list_to_row(self, row_to_add=[], auto_move=True):
        for offset, val in enumerate(row_to_add):
            cell = self.worksheet.cell(column=self.cursor[0] + offset, row=self.cursor[1])
            cell.value = val
        if auto_move:
            self.shift_down()

    def write_list_to_column(self, column_to_add=[], auto_move=True):
        for offset, val in enumerate(column_to_add):
            cell = self.worksheet.cell(column=self.cursor[0], row=self.cursor[1] + offset)
            cell.value = val
        if auto_move:
            self.shift_down(ammount=len(column_to_add))

    def write_2dlist_to_table(self, list_of_lists=[], auto_move=True):
        for y_offset, row in enumerate(list_of_lists):
            for x_offset, val in enumerate(row):
                cell = self.worksheet.cell(column=self.cursor[0] + x_offset, row=self.cursor[1] + y_offset)
                cell.value = val
        if auto_move:
            self.shift_down(ammount=len(list_of_lists))

    def __enter__(self):
        self.workbook = load_workbook(filename=self.filename)
        self.worksheet = self.workbook.worksheets[0]
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.workbook.save(filename=self.filename)
